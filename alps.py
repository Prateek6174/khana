import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import datetime


st.set_page_config(
    page_title="Food Preference Manager",
    page_icon="🍽️",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ── Session state ──────────────────────────────────────────────────────────────
for k, v in [("df", None), ("success", None), ("swiss", True), ("lang", "EN")]:
    if k not in st.session_state:
        st.session_state[k] = v

swiss    = st.session_state.swiss
lang_key = st.session_state.lang if swiss else "EN"

LANGS = {
    "EN": {
        "eyebrow": "Confidential-To be accessed only by the GC Lead Admin ",
        "title": "TATA's Bread and Butter",
        "sub": "From the streets of Kota to the quiet hills of Lenzburg.",
        "day_label": "SELECT A DAY TO GENERATE ALL 8 SHEETS",
        "days": ["Monday","Tuesday","Wednesday","Thursday","Friday"],
        "loading": "Generating sheets for {}…",
        "saved": "8 sheets saved",
        "tagline": "",
    },
    "DE": {
        "eyebrow": "Vertraulich – Zugriff nur für den GC Lead Admin",
        "title": "Tatas Brot und Butter",
        "sub": "Von den Straßen Kotas bis zu den stillen Hügeln Lenzburgs.",
        "day_label": "TAG AUSWÄHLEN — 7 TABELLEN WERDEN ERSTELLT",
        "days": ["Montag","Dienstag","Mittwoch","Donnerstag","Freitag"],
        "loading": "Schweizer Präzision am Werk für {}…",
        "saved": "7 Tabellen gespeichert",
        "tagline": "",
    },
    "FR": {
        "eyebrow": "Confidentiel – Accès réservé à l’administrateur principal du GC",
        "title": "Pain et beurre TATA",
        "sub": "Des rues de Kota aux collines paisibles de Lenzburg.",
        "day_label": "SÉLECTIONNEZ UN JOUR",
        "days": ["Lundi","Mardi","Mercredi","Jeudi","Vendredi"],
        "loading": "Précision suisse en action pour {}…",
        "saved": "7 fichiers enregistrés",
        "tagline": "",
    },
    "IT": {
        "eyebrow": "Riservato - Accesso consentito solo all'amministratore principale di GC",
        "title": "Il pane e il burro di TATA",
        "sub": "Dalle strade di Kota alle tranquille colline di Lenzburg.",
        "day_label": "SELEZIONA UN GIORNO",
        "days": ["Lunedì","Martedì","Mercoledì","Giovedì","Venerdì"],
        "loading": "Precisione svizzera in azione per {}…",
        "saved": "7 fogli salvati",
        "tagline": "",
    },
}
T = LANGS[lang_key]

DAY_MAP_EN = {
    "Monday":"mon","Tuesday":"tue","Wednesday":"wed","Thursday":"thu","Friday":"fri",
    "Montag":"mon","Dienstag":"tue","Mittwoch":"wed","Donnerstag":"thu","Freitag":"fri",
    "Lundi":"mon","Mardi":"tue","Mercredi":"wed","Jeudi":"thu","Vendredi":"fri",
    "Lunedì":"mon","Martedì":"tue","Mercoledì":"wed","Giovedì":"thu","Venerdì":"fri",
}
DAY_EN_MAP = {
    "Monday":"Monday","Tuesday":"Tuesday","Wednesday":"Wednesday","Thursday":"Thursday","Friday":"Friday",
    "Montag":"Monday","Dienstag":"Tuesday","Mittwoch":"Wednesday","Donnerstag":"Thursday","Freitag":"Friday",
    "Lundi":"Monday","Mardi":"Tuesday","Mercredi":"Wednesday","Jeudi":"Thursday","Vendredi":"Friday",
    "Lunedì":"Monday","Martedì":"Tuesday","Mercoledì":"Wednesday","Giovedì":"Thursday","Venerdì":"Friday",
}
CANTON_NAMES = ["Zürich","Bern","Genève","Basel","Luzern"]
SWISS_HOLIDAYS = [
    (1,1,"Neujahr / New Year's Day"),(1,2,"Berchtoldstag"),
    (4,3,"Karfreitag / Good Friday"),(4,6,"Ostermontag / Easter Monday"),
    (5,1,"Tag der Arbeit / Labour Day"),(5,14,"Auffahrt / Ascension Day"),
    (5,25,"Pfingstmontag / Whit Monday"),(8,1,"🎉 Bundesfeiertag / Swiss National Day"),
    (12,25,"Weihnachten / Christmas Day"),(12,26,"Stephanstag / St. Stephen's Day"),
]

# ══════════════════════════════════════════════════════════════════════════════
#  DESIGN TOKENS
# ══════════════════════════════════════════════════════════════════════════════
if swiss:
    # Swiss mode — rich red flag background, bright white cross tiles, light card
    bg_html      = "#C8102E"   # Swiss flag red (exact pantone 485 C)
    # Cross pattern SVG with BRIGHT WHITE crosses (fully visible)
    cross_svg_url = (
        "url(\"data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg'"
        " width='140' height='140'%3E"
        "%3Crect width='140' height='140' fill='%23C8102E'/%3E"
        # Vertical bar — pure white
        "%3Crect x='60' y='24' width='20' height='92' rx='4' fill='white' opacity='0.22'/%3E"
        # Horizontal bar — pure white
        "%3Crect x='24' y='60' width='92' height='20' rx='4' fill='white' opacity='0.22'/%3E"
        "%3C/svg%3E\")"
    )
    card_bg      = "rgba(255,255,255,0.13)"
    card_border  = "rgba(255,255,255,0.28)"
    card_shadow  = "0 24px 64px rgba(120,0,0,0.45), 0 4px 16px rgba(0,0,0,0.25)"
    font_family  = "'Helvetica Neue','Arial',sans-serif"
    accent       = "#ffffff"
    accent_dim   = "rgba(255,255,255,0.6)"
    accent_glow  = "rgba(255,255,255,0.18)"
    text_main    = "#ffffff"           # Pure white — max legibility on dark card
    text_muted   = "rgba(255,255,255,0.65)"
    text_body    = "#000000"           # Warm white for body text
    border_col   = "rgba(255,255,255,0.2)"
    btn_bg       = "rgba(255,255,255,0.10)"
    btn_hover    = "rgba(255,255,255,0.22)"
    divider_col  = "rgba(255,255,255,0.12)"
    success_bg   = "rgba(255,255,255,0.10)"
    success_bdr  = "rgba(255,255,255,0.3)"
    success_col  = "#ffffff"
    input_bg     = "rgba(255,255,255,0.07)"
    upload_border= "rgba(255,255,255,0.3)"
else:
    # Normal mode — light blue, dark readable text
    bg_html      = "#dce8f5"   # Soft sky blue
    card_bg      = "#ffffff"
    card_border  = "rgba(37,99,235,0.12)"
    card_shadow  = "0 8px 40px rgba(37,99,235,0.10), 0 2px 8px rgba(0,0,0,0.06)"
    cross_svg_url= "none"
    font_family  = "'Inter',sans-serif"
    accent       = "#1d4ed8"   # Deep blue accent
    accent_dim   = "rgba(29,78,216,0.6)"
    accent_glow  = "rgba(29,78,216,0.12)"
    text_main    = "#0f172a"   # Near-black — crisp on white card
    text_muted   = "#475569"   # Slate — comfortable on white
    text_body    = "#334155"
    border_col   = "rgba(29,78,216,0.14)"
    btn_bg       = "#eff6ff"   # Very light blue
    btn_hover    = "#dbeafe"
    divider_col  = "rgba(29,78,216,0.1)"
    success_bg   = "rgba(21,128,61,0.07)"
    success_bdr  = "rgba(21,128,61,0.25)"
    success_col  = "#15803d"
    input_bg     = "#f8faff"
    upload_border= "rgba(29,78,216,0.25)"

# Snowfall JS — active only in swiss mode
snowfall_js = """
<canvas id="snow" style="position:fixed;top:0;left:0;width:100%;height:100%;
    pointer-events:none;z-index:9999;"></canvas>
<script>
(function(){
  const c = document.getElementById('snow');
  const ctx = c.getContext('2d');
  let W = window.innerWidth, H = window.innerHeight;
  c.width = W; c.height = H;
  window.addEventListener('resize', ()=>{ W=c.width=window.innerWidth; H=c.height=window.innerHeight; });

  const N = 160;
  const flakes = Array.from({length:N}, ()=>({
    x: Math.random()*W,
    y: Math.random()*H,
    r: Math.random()*3.5 + 1.2,
    sx: (Math.random()-0.5)*0.6,
    sy: Math.random()*1.2 + 0.4,
    o: Math.random()*0.5 + 0.3,
    wobble: Math.random()*Math.PI*2,
    wobbleSpeed: Math.random()*0.02 + 0.005,
  }));

  function draw(){
    ctx.clearRect(0,0,W,H);
    flakes.forEach(f=>{
      ctx.beginPath();
      ctx.arc(f.x, f.y, f.r, 0, Math.PI*2);
      ctx.fillStyle = `rgba(255,255,255,${f.o})`;
      ctx.shadowBlur = 4;
      ctx.shadowColor = 'rgba(255,255,255,0.6)';
      ctx.fill();

      f.wobble += f.wobbleSpeed;
      f.x += f.sx + Math.sin(f.wobble)*0.4;
      f.y += f.sy;

      if(f.y > H+10){
        f.y = -10;
        f.x = Math.random()*W;
      }
      if(f.x > W+10) f.x = -10;
      if(f.x < -10)  f.x = W+10;
    });
    requestAnimationFrame(draw);
  }
  draw();
})();
</script>
""" if swiss else ""

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

html, body, [class*="css"] {{
    font-family: {font_family} !important;
}}

/* ── FULL VIEWPORT BACKGROUND ── */
html {{ background: {bg_html} !important; }}
body {{ background: {bg_html} !important; }}
.stApp {{
    background: {cross_svg_url} repeat, {bg_html} !important;
    min-height: 100vh;
}}
[data-testid="stAppViewContainer"] {{
    background: {cross_svg_url} repeat, {bg_html} !important;
}}

/* ── CENTRE CARD ── */
.block-container {{
    position: relative;
    z-index: 1;
    max-width: 760px !important;
    padding: 2.6rem 3rem 4rem !important;
    margin-top: 1rem !important;
    background: {card_bg} !important;
    {"backdrop-filter: blur(24px) saturate(1.6) !important;" if swiss else ""}
    {"-webkit-backdrop-filter: blur(24px) saturate(1.6) !important;" if swiss else ""}
    border-radius: 22px !important;
    border: 1px solid {card_border} !important;
    box-shadow: {card_shadow} !important;
}}

#MainMenu{{visibility:hidden;}} footer{{visibility:hidden;}} header{{visibility:hidden;}}

/* ── CORNER BUTTONS — fixed outside the centre card ── */
.corner-nav-left,
.corner-nav-right {{
    position: fixed;
    top: 1.1rem;
    z-index: 10000;
    pointer-events: all;
}}
.corner-nav-left  {{ left:  1.4rem; }}
.corner-nav-right {{ right: 1.4rem; }}
.corner-nav-left  a,
.corner-nav-right a {{
    display: inline-flex; align-items: center; gap: 0.3rem;
    font-family: {font_family}; font-size: 0.74rem; font-weight: 700;
    letter-spacing: 0.8px; padding: 0.44rem 1rem;
    border-radius: 22px; cursor: pointer; text-decoration: none;
    transition: all 0.18s ease;
    background: {"rgba(255,255,255,0.18)" if swiss else "#1d4ed8"};
    color: #ffffff;
    border: 1px solid {"rgba(255,255,255,0.38)" if swiss else "transparent"};
    box-shadow: {"0 2px 16px rgba(0,0,0,0.3)" if swiss else "0 2px 14px rgba(29,78,216,0.4)"};
}}
.corner-nav-left  a:hover,
.corner-nav-right a:hover {{
    transform: translateY(-1px);
    box-shadow: {"0 4px 20px rgba(0,0,0,0.4)" if swiss else "0 4px 20px rgba(29,78,216,0.5)"};
    background: {"rgba(255,255,255,0.28)" if swiss else "#1e40af"};
}}

/* ── HERO ── */
.hero {{ text-align:center; margin-bottom:2.2rem; }}
.hero-eyebrow {{
    display:inline-block; font-size:0.67rem; font-weight:700;
    letter-spacing:2.8px; text-transform:uppercase;
    color:{accent_dim}; margin-bottom:0.85rem;
}}
.hero-title {{
    font-size:1.95rem; font-weight:700; color:{text_main} !important;
    margin:0 0 0.5rem; letter-spacing:-0.5px; line-height:1.2;
}}
h1, h2, h3 {{
    color: {text_main} !important;
}}
.hero-sub {{ font-size:0.9rem; color:{text_muted}; margin:0; }}
.hero-tagline {{
    margin-top:0.65rem; font-size:0.7rem; font-weight:700;
    letter-spacing:2.5px; text-transform:uppercase; color:{accent_dim};
}}

/* ── Upload ── */
[data-testid="stFileUploaderDropzone"] {{
    background: {input_bg} !important;
    border: 1.5px dashed {upload_border} !important;
    border-radius: 14px !important;
    transition: all 0.2s ease !important;
}}
[data-testid="stFileUploaderDropzone"]:hover {{
    border-color: {accent} !important;
    background: {accent_glow} !important;
}}

/* ── Divider ── */
.divider {{ height:1px; background:{divider_col}; margin:2.2rem 0; }}

/* ── Day label ── */
.day-label {{
    text-align:center; font-size:0.67rem; font-weight:700;
    letter-spacing:2.8px; text-transform:uppercase;
    color:{accent_dim}; margin-bottom:1.1rem;
}}

/* ── Day buttons ── */
.stButton > button {{
    background: {btn_bg} !important;
    color: {text_muted} !important;
    border: 1px solid {border_col} !important;
    border-radius: 12px !important;
    height: 76px !important; width: 100% !important;
    font-family: {font_family} !important;
    font-size: 0.81rem !important; font-weight: 600 !important;
    letter-spacing: 0.3px !important;
    transition: all 0.18s ease !important;
    {"color: rgba(255,255,255,0.7) !important;" if swiss else f"color: {text_muted} !important;"}
}}
.stButton > button:hover {{
    background: {btn_hover} !important;
    border-color: {accent} !important;
    color: {"#ffffff" if swiss else text_main} !important;
    box-shadow: 0 4px 22px {accent_glow} !important;
    transform: translateY(-2px) !important;
}}
.stButton > button:active {{ transform: translateY(0) !important; }}

/* ── Toast ── */
.toast {{
    margin-top:2rem; display:flex; align-items:center;
    justify-content:center; gap:0.6rem;
    background:{success_bg}; border:1px solid {success_bdr};
    border-radius:40px; padding:0.75rem 1.6rem;
    color:{success_col}; font-size:0.85rem; font-weight:500;
}}

/* ── Holiday banner ── */
.holiday-banner {{
    background: {"rgba(255,255,255,0.10)" if swiss else "rgba(29,78,216,0.06)"};
    border: 1px solid {"rgba(255,255,255,0.22)" if swiss else "rgba(29,78,216,0.18)"};
    border-radius:12px; padding:0.6rem 1.2rem;
    color:{"rgba(255,255,255,0.88)" if swiss else "#1d4ed8"};
    font-size:0.8rem; text-align:center; margin-bottom:1.2rem;
}}

/* ── Clock ── */
.cet-clock {{
    text-align:right; margin-top:0.2rem; margin-bottom:1rem;
}}

.stSpinner > div {{ border-top-color:{accent} !important; }}
label {{ display:none !important; }}
/* All text white in swiss mode */
p, li, span, div {{ color:{text_body} !important; }}
h1, h2, h3, h4, h5, h6 {{ color:{text_main} !important; }}
[data-testid="stMarkdownContainer"],
[data-testid="stMarkdownContainer"] * {{ color:{text_body} !important; }}
[data-testid="stMarkdownContainer"] h1,
[data-testid="stMarkdownContainer"] h2,
[data-testid="stMarkdownContainer"] h3 {{ color:{text_main} !important; }}
[data-testid="stFileUploaderDropzone"] *,
[data-testid="stFileUploaderDropzone"] span {{ color:{text_body} !important; }}
.stSpinner > div > div {{ color:{text_body} !important; }}
</style>

{snowfall_js}

<!-- Corner nav injected below via separate st.markdown -->


""", unsafe_allow_html=True)

# ── Corner nav buttons — separate markdown so no f-string escaping issues ────
# ── Corner nav buttons — same-window Streamlit buttons ────
# ── Corner nav buttons — same-window Streamlit buttons ───────────────────────

# ── Query-param based navigation (corner buttons) ────────────────────────────)

# ── Language switcher (Swiss only, horizontal pills) ──────────────────────────
if swiss:
    lang_labels = {"EN": "English", "DE": "Deutsch", "FR": "Français", "IT": "Italiano"}
    lc1,lc2,lc3,lc4,_ = st.columns([1.8,1.7,1.9,1.7,1.4], gap="small")
    for col_l, lk in zip([lc1,lc2,lc3,lc4], ["EN","DE","FR","IT"]):
        with col_l:
            mark = "✦ " if st.session_state.lang == lk else ""
            if st.button(f"{mark}{lang_labels[lk]}", key=f"lang_{lk}", use_container_width=True):
                st.session_state.lang = lk; st.rerun()

# ── CET Clock (Swiss only) ─────────────────────────────────────────────────────
if swiss:
    try:
        import pytz
        now_ch = datetime.datetime.now(pytz.timezone("Europe/Zurich"))
    except Exception:
        now_ch = datetime.datetime.utcnow() + datetime.timedelta(hours=2)
    _, clk_col = st.columns([3,1])
    with clk_col:
        st.markdown(f"""
        <div class="cet-clock">
            <div style="font-size:1.1rem;font-weight:700;letter-spacing:2px;
                color:rgba(255,255,255,0.9);font-variant-numeric:tabular-nums;">
                {now_ch.strftime("%H:%M:%S")}
            </div>
            <div style="font-size:0.62rem;color:rgba(255,255,255,0.4);
                letter-spacing:1px;margin-top:1px;">
                {now_ch.strftime("%a %d %b")} · CET
            </div>
        </div>
        """, unsafe_allow_html=True)

    # Holiday check
    holiday = next((n for m,d,n in SWISS_HOLIDAYS if m==now_ch.month and d==now_ch.day), None)
    if holiday:
        st.markdown(f'<div class="holiday-banner">🇨🇭 &nbsp; Swiss Public Holiday: <strong>{holiday}</strong></div>',
                    unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
#  HERO
# ══════════════════════════════════════════════════════════════════════════════
if swiss:
    st.markdown(f"""
    <div class="hero" style="padding-top:0.5rem;">
        <!-- Swiss flag SVG with glow -->
        <div style="display:flex;justify-content:center;margin-bottom:1.4rem;">
            <svg width="68" height="68" viewBox="0 0 68 68" xmlns="http://www.w3.org/2000/svg"
                 style="filter:drop-shadow(0 0 18px rgba(255,255,255,0.35))
                        drop-shadow(0 4px 24px rgba(200,16,46,0.7));">
                <rect width="68" height="68" rx="8" fill="#C8102E"/>
                <rect x="27" y="11" width="14" height="46" rx="3" fill="white"/>
                <rect x="11" y="27" width="46" height="14" rx="3" fill="white"/>
            </svg>
        </div>
        <span class="hero-eyebrow">🍽️ &nbsp; {T["eyebrow"]}</span>
        <h1 class="hero-title">{T["title"]}</h1>
        <p class="hero-sub" style="color:rgba(255,255,255,0.65);">{T["sub"]}</p>
        <div class="hero-tagline">{T["tagline"]}</div>
        <div style="margin-top:1rem;font-size:0.82rem;letter-spacing:4px;color:rgba(255,255,255,0.2);">
            🧀 &nbsp; 🍫 &nbsp; ⛷️ &nbsp; 🏔️ &nbsp; ⌚ &nbsp; 🪗
        </div>
        <!-- Alps silhouette -->
        <div style="margin-top:1.3rem;opacity:0.18;">
            <svg viewBox="0 0 700 55" xmlns="http://www.w3.org/2000/svg" style="width:100%;display:block;">
                <path d="M0,55 L55,34 L100,46 L170,14 L230,32 L285,4 L340,20
                         L370,0 L400,18 L455,8 L510,28 L565,12 L620,34 L670,20
                         L700,30 L700,55 Z" fill="white"/>
            </svg>
        </div>
    </div>
    """, unsafe_allow_html=True)
else:
    st.markdown(f"""
    <div class="hero" style="padding-top:1rem;">
        <span class="hero-eyebrow">🍽️ &nbsp; {T["eyebrow"]}</span>
        <h1 class="hero-title">{T["title"]}</h1>
        <p class="hero-sub">{T["sub"]}</p>
    </div>
    """, unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL ENGINE
# ══════════════════════════════════════════════════════════════════════════════
CATEGORY_COLORS = {
    "Breakfast - Eggs":    ("1D4ED8","DBEAFE"),
    "Breakfast - No Eggs": ("1E40AF","EFF6FF"),
    "Lunch - Veg":         ("15803D","DCFCE7"),
    "Lunch - Non-Veg":     ("B45309","FEF3C7"),
    "Lunch - Jain":        ("7E22CE","F5F3FF"),
    "Lunch - Fruit Plate": ("065F46","D1FAE5"),
    "Snacks":              ("0E7490","E0F2FE"),
}

def create_excel(category_name, employees):
    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = category_name[:31]; ws.sheet_view.showGridLines = False
    td, tl = CATEGORY_COLORS.get(category_name, ("1D4ED8","DBEAFE"))
    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.row_dimensions[1].height = 40
    ws.merge_cells("A1:C1"); t = ws["A1"]
    t.value = category_name.upper()
    t.font  = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    t.fill  = PatternFill("solid", fgColor=td)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 28
    for cl, hdr, w in [("A","Employee Name",36),("B","Employee ID",18),("C","Signature",32)]:
        c = ws[f"{cl}2"]; c.value = hdr
        c.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1E3A5F")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bdr; ws.column_dimensions[cl].width = w
    for idx, emp in enumerate(employees):
        r = 3+idx; ws.row_dimensions[r].height = 22
        bg = tl if idx%2==0 else "FFFFFF"
        cn = ws.cell(r,1,str(emp["name"]).strip()); cn.font=Font(name="Calibri",size=11,color="1E293B"); cn.fill=PatternFill("solid",fgColor=bg); cn.alignment=Alignment(horizontal="left",vertical="center",indent=1); cn.border=bdr
        ci = ws.cell(r,2,str(emp["id"]).strip());   ci.font=Font(name="Calibri",size=11,color="1E293B"); ci.fill=PatternFill("solid",fgColor=bg); ci.alignment=Alignment(horizontal="center",vertical="center"); ci.border=bdr
        cs = ws.cell(r,3,"");                        cs.font=Font(name="Calibri",size=11); cs.fill=PatternFill("solid",fgColor=bg); cs.border=bdr
    fr = 3+len(employees)+1; ws.row_dimensions[fr].height=16
    ws.merge_cells(f"A{fr}:C{fr}"); fc=ws[f"A{fr}"]
    fc.value=f"Total Employees: {len(employees)}"; fc.font=Font(name="Calibri",italic=True,size=9,color="94A3B8"); fc.alignment=Alignment(horizontal="right",vertical="center")
    ws.print_title_rows="4:4"; return wb

def create_summary_excel(day_en, summary_rows, total_unique, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14

    title_fill = PatternFill("solid", fgColor="7C1D1D")
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    body_fill = PatternFill("solid", fgColor="F8FAFC")
    accent_fill = PatternFill("solid", fgColor="DBEAFE")
    total_fill = PatternFill("solid", fgColor="DCFCE7")

    thin = Side(style="thin", color="D1D5DB")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("B2:E2")
    c = ws["B2"]
    c.value = f"Meal Summary — {day_en}"
    c.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    c.fill = title_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 28

    ws.merge_cells("B3:E3")
    c = ws["B3"]
    c.value = "Unique employees counted once by Employee ID. Duplicate IDs are removed before summary generation."
    c.font = Font(name="Calibri", size=10, italic=True, color="475569")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 34

    headers = ["Category", "Choice", "Count", "Share %"]
    for col, hdr in enumerate(headers, start=2):
        cell = ws.cell(5, col, hdr)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = bdr

    row = 6
    for section, choice, count in summary_rows:
        share = (count / total_unique) if total_unique else 0
        values = [section, choice, count, share]

        for col, val in enumerate(values, start=2):
            cell = ws.cell(row, col, val)
            cell.border = bdr
            cell.alignment = Alignment(
                horizontal="center" if col in (4, 5) else "left",
                vertical="center"
            )
            cell.font = Font(name="Calibri", size=11, bold=(col == 2))
            cell.fill = accent_fill if col == 2 else body_fill
            if col == 5:
                cell.number_format = "0.0%"
        row += 1

    ws.merge_cells(f"B{row}:C{row}")
    c1 = ws[f"B{row}"]
    c1.value = "Total Unique Employees"
    c1.font = Font(name="Calibri", size=11, bold=True, color="166534")
    c1.fill = total_fill
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c1.border = bdr

    c2 = ws[f"D{row}"]
    c2.value = total_unique
    c2.font = Font(name="Calibri", size=11, bold=True, color="166534")
    c2.fill = total_fill
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c2.border = bdr

    c3 = ws[f"E{row}"]
    c3.value = 1 if total_unique else 0
    c3.number_format = "0.0%"
    c3.font = Font(name="Calibri", size=11, bold=True, color="166534")
    c3.fill = total_fill
    c3.alignment = Alignment(horizontal="center", vertical="center")
    c3.border = bdr

    row += 2
    ws.merge_cells(f"B{row}:E{row}")
    c = ws[f"B{row}"]
    c.value = f"Generated on {datetime.datetime.now().strftime('%d %b %Y %H:%M')}"
    c.font = Font(name="Calibri", size=9, italic=True, color="64748B")
    c.alignment = Alignment(horizontal="right", vertical="center")

    wb.save(str(output_path))

def generate_excels_for_day(df, day_local, output_base):
    day_key = DAY_MAP_EN.get(day_local, "mon")
    day_en = DAY_EN_MAP.get(day_local, day_local)

    d = Path(output_base) / day_en
    bf = d / "Breakfast"
    lf = d / "Lunch"
    sf = d / "Snacks"
    sm = d / "Summary"

    for folder in (bf, lf, sf, sm):
        folder.mkdir(parents=True, exist_ok=True)

    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    nc = ic = bc = lc = sc = None
    for c in df.columns:
        cl = str(c).lower().strip().replace("\n", " ").replace("\r", " ")

        if cl == "employee name":
            nc = c
        if cl in ("employee id", "employee"):
            ic = c
        if "breakfast" in cl and day_key in cl:
            bc = c
        if "lunch" in cl and day_key in cl:
            lc = c
        if "snack" in cl:
            sc = c

    if not nc:
        raise ValueError(f"Employee Name column not found. Columns: {list(df.columns)}")
    if not ic:
        raise ValueError(f"Employee ID / Employee column not found. Columns: {list(df.columns)}")

    def clean_id(v):
        if pd.isna(v):
            return ""
        s = str(v).strip()
        if not s or s.lower() == "nan":
            return ""
        try:
            return str(int(float(s)))
        except Exception:
            return s

    def clean_name(v):
        if pd.isna(v):
            return ""
        s = str(v).strip()
        return "" if (not s or s.lower() == "nan") else s

    def norm(v):
        if pd.isna(v):
            return ""
        s = str(v).strip().lower()
        s = s.replace("_", " ").replace("-", " ")
        s = " ".join(s.split())
        return s

    df[ic] = df[ic].apply(clean_id)
    df[nc] = df[nc].apply(clean_name)

    df = df[(df[ic] != "") & (df[nc] != "")].copy()

    order_cols = [
        c for c in df.columns
        if str(c).strip().lower() in ("completion time", "timestamp", "completed at", "end time")
    ]

    if order_cols:
        oc = order_cols[0]
        try:
            dt = pd.to_datetime(df[oc], errors="coerce")
            df = df.assign(__order__=dt, __seq__=range(len(df))).sort_values(["__order__", "__seq__"])
        except Exception:
            df = df.assign(__seq__=range(len(df))).sort_values(["__seq__"])
    else:
        df = df.assign(__seq__=range(len(df))).sort_values(["__seq__"])

    df = df.drop_duplicates(subset=[ic], keep="last").reset_index(drop=True)

    false_mask = pd.Series(False, index=df.index)

    def get_people(mask):
        subset = df.loc[mask, [nc, ic]].copy()
        subset = subset.drop_duplicates(subset=[ic], keep="first")
        return [
            {"name": str(r[nc]).strip(), "id": str(r[ic]).strip()}
            for _, r in subset.iterrows()
            if str(r[nc]).strip() and str(r[ic]).strip()
        ]

    bnorm = df[bc].apply(norm) if bc else pd.Series([""] * len(df), index=df.index)
    lnorm = df[lc].apply(norm) if lc else pd.Series([""] * len(df), index=df.index)

    breakfast_people = {
        "Eggs": get_people(bnorm.eq("with eggs") if bc else false_mask),
        "No Eggs": get_people((bnorm.eq("without eggs") | bnorm.eq("no eggs")) if bc else false_mask),
    }

    lunch_people = {
        "Veg": get_people(lnorm.eq("veg") if lc else false_mask),
        "Non-Veg": get_people((lnorm.eq("non veg") | lnorm.eq("nonveg")) if lc else false_mask),
        "Jain": get_people(lnorm.eq("jain") if lc else false_mask),
        "Fruit Plate": get_people((lnorm.eq("fruit plate") | lnorm.eq("fruitplate")) if lc else false_mask),
    }

    if sc:
        def snack_match(x):
            if pd.isna(x):
                return False
            s = str(x).strip()
            if not s or s.lower() == "nan":
                return False
            tokens = [t.strip().lower() for t in s.split(";") if t.strip()]
            return day_key.lower() in tokens

        snack_people = get_people(df[sc].apply(snack_match))
    else:
        snack_people = []

    create_excel("Breakfast - Eggs", breakfast_people["Eggs"]).save(str(bf / "Breakfast_Eggs.xlsx"))
    create_excel("Breakfast - No Eggs", breakfast_people["No Eggs"]).save(str(bf / "Breakfast_No_Eggs.xlsx"))

    create_excel("Lunch - Veg", lunch_people["Veg"]).save(str(lf / "Lunch_Veg.xlsx"))
    create_excel("Lunch - Non-Veg", lunch_people["Non-Veg"]).save(str(lf / "Lunch_NonVeg.xlsx"))
    create_excel("Lunch - Jain", lunch_people["Jain"]).save(str(lf / "Lunch_Jain.xlsx"))
    create_excel("Lunch - Fruit Plate", lunch_people["Fruit Plate"]).save(str(lf / "Lunch_Fruit_Plate.xlsx"))

    create_excel("Snacks", snack_people).save(str(sf / f"Snacks_{day_en}.xlsx"))

    summary_rows = [
        ("Breakfast", "Eggs", len(breakfast_people["Eggs"])),
        ("Breakfast", "No Eggs", len(breakfast_people["No Eggs"])),
        ("Lunch", "Veg", len(lunch_people["Veg"])),
        ("Lunch", "Non-Veg", len(lunch_people["Non-Veg"])),
        ("Lunch", "Jain", len(lunch_people["Jain"])),
        ("Lunch", "Fruit Plate", len(lunch_people["Fruit Plate"])),
        ("Snacks", day_en, len(snack_people)),
    ]

    total_unique = int(df[ic].nunique())
    create_summary_excel(day_en, summary_rows, total_unique, sm / f"Summary_{day_en}.xlsx")
# ══════════════════════════════════════════════════════════════════════════════
#  MAIN CONTENT
# ══════════════════════════════════════════════════════════════════════════════

# Upload
uploaded = st.file_uploader(" ", type=["xlsx","xls"])
if uploaded:
    try:
        df_loaded = pd.read_excel(uploaded, engine="openpyxl")
        df_loaded.columns = [str(c).strip() for c in df_loaded.columns]
        st.session_state.df      = df_loaded
        st.session_state.success = None
    except Exception as e:
        st.error(f"Could not read file: {e}")

# Day buttons
if st.session_state.df is not None:
    st.markdown('<div class="divider"></div>', unsafe_allow_html=True)
    st.markdown(f'<p class="day-label">{T["day_label"]}</p>', unsafe_allow_html=True)
    output_path = str(Path.home() / "Downloads" / "Khaana")
    days   = T["days"]
    dcols  = st.columns(5, gap="small")
    clicked = None
    for i, (col, day) in enumerate(zip(dcols, days)):
        with col:
            lbl = f"{day}\n{CANTON_NAMES[i]}" if swiss else day
            if st.button(lbl, key=f"btn_{day}", use_container_width=True):
                clicked = day

    if clicked:
        with st.spinner(T["loading"].format(clicked)):
            try:
                generate_excels_for_day(st.session_state.df, clicked, output_path)
                st.session_state.success = clicked
                if swiss: st.balloons()
            except Exception as e:
                st.error(f"Error: {e}")
                import traceback; st.code(traceback.format_exc())

    if st.session_state.success:
        day_en   = DAY_EN_MAP.get(st.session_state.success, st.session_state.success)
        day_path = Path(output_path) / day_en
        if swiss:
            st.markdown(f"""
            <div class="toast">
                <svg width="15" height="15" viewBox="0 0 15 15" xmlns="http://www.w3.org/2000/svg" style="flex-shrink:0;">
                    <rect width="15" height="15" rx="2.5" fill="#C8102E" stroke="rgba(255,255,255,0.5)" stroke-width="1"/>
                    <rect x="6" y="2.5" width="3" height="10" rx="1" fill="white"/>
                    <rect x="2.5" y="6" width="10" height="3" rx="1" fill="white"/>
                </svg>
                &nbsp; {T["saved"]} &nbsp;·&nbsp; Schweizer Präzision ✓ &nbsp;·&nbsp;
                <code style="color:rgba(255,255,255,0.65);font-size:0.76rem;">{day_path}</code>
            </div>
            """, unsafe_allow_html=True)
        else:
            st.markdown(f"""
            <div class="toast">
                ✅ &nbsp; {T["saved"]} &nbsp;·&nbsp;
                <code style="font-size:0.8rem;">{day_path}</code>
            </div>
            """, unsafe_allow_html=True)

# Swiss footer
if swiss:
    st.markdown("""
    <div style="text-align:center;margin-top:3.5rem;padding-top:1.8rem;
        border-top:1px solid rgba(255,255,255,0.1);">
        <svg width="26" height="26" viewBox="0 0 26 26" xmlns="http://www.w3.org/2000/svg"
             style="opacity:0.3;margin-bottom:0.5rem;">
            <rect width="26" height="26" rx="4" fill="#C8102E" stroke="rgba(255,255,255,0.4)" stroke-width="1"/>
            <rect x="10.5" y="4" width="5" height="18" rx="1.5" fill="white"/>
            <rect x="4" y="10.5" width="18" height="5" rx="1.5" fill="white"/>
        </svg>
        <div style="font-size:0.62rem;font-weight:700;letter-spacing:2.5px;
            text-transform:uppercase;color:rgba(255,255,255,0.25);">
            Made in India &nbsp;·&nbsp; For the people of Switzerland
        </div>
        <div style="font-size:0.57rem;color:rgba(255,255,255,0.13);
            margin-top:0.3rem;letter-spacing:0.8px;">
            Liberté · Gleichheit · Libertà · Freedom
        </div>
    </div>
    """, unsafe_allow_html=True)
